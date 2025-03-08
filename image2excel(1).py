# @Version  : 1.0
# @Author   : QiHack
# @Github   :https://github.com/zzqyyds365


import sys
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image
from openpyxl.utils import get_column_letter

def draw_image_in_excel(image_path, output_path, max_width=300, max_height=300):
    # 打开图片
    image = Image.open(image_path)

    # 调整图片大小
    width, height = image.size
    if width > max_width or height > max_height:
        ratio = min(max_width / width, max_height / height)
        new_width = int(width * ratio)
        new_height = int(height * ratio)
        image = image.resize((new_width, new_height), Image.LANCZOS)

    # 创建 Excel 工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 遍历图片像素
    width, height = image.size
    for y in range(height):
        for x in range(width):
            # 获取像素的 RGB 颜色值
            r, g, b = image.getpixel((x, y))[:3]
            # 将 RGB 颜色值转换为十六进制字符串
            color_hex = '{:02x}{:02x}{:02x}'.format(r, g, b)
            # 创建填充样式
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            # 设置单元格的填充样式
            ws.cell(row=y + 1, column=x + 1).fill = fill

    # 设置列宽，这里将列宽设置为 2，你可以根据需要调整这个值
    default_column_width = 2
    for col in range(1, width + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = default_column_width

    # 保存 Excel 文件
    wb.save(output_path)
    result_label.config(text=f"图片已成功绘制到 Excel 文件：{output_path}")

def select_image():
    image_path = filedialog.askopenfilename(filetypes=[("图片文件", "*.jpg;*.png")])
    image_entry.delete(0, tk.END)
    image_entry.insert(0, image_path)

def select_output():
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
    output_entry.delete(0, tk.END)
    output_entry.insert(0, output_path)

def convert_image():
    image_path = image_entry.get()
    output_path = output_entry.get()
    if image_path and output_path:
        draw_image_in_excel(image_path, output_path)
    else:
        result_label.config(text="请选择图片和输出文件路径！")

# 创建主窗口
root = tk.Tk()
root.title("图片转 Excel")

# 图片选择部分
image_label = tk.Label(root, text="选择图片:")
image_label.pack(pady=5)
image_entry = tk.Entry(root, width=50)
image_entry.pack(pady=5)
image_button = tk.Button(root, text="浏览", command=select_image)
image_button.pack(pady=5)

# 输出文件选择部分
output_label = tk.Label(root, text="选择输出文件:")
output_label.pack(pady=5)
output_entry = tk.Entry(root, width=50)
output_entry.pack(pady=5)
output_button = tk.Button(root, text="浏览", command=select_output)
output_button.pack(pady=5)

# 转换按钮
convert_button = tk.Button(root, text="转换", command=convert_image)
convert_button.pack(pady=20)

# 结果显示标签
result_label = tk.Label(root, text="")
result_label.pack(pady=10)

# 运行主循环
root.mainloop()